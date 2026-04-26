"""Validate that an uploaded workbook matches the canonical CIE+SEE template.

The parser is forgiving — too forgiving — and silently produces wrong numbers
when the layout drifts (missing CO tag, missing max marks, mis-rowed labels,
non-int Sl.No mid-range, etc.). This module checks the structural contract
*before* parsing and returns a structured list of violations.

    violations == []  → workbook conforms; safe to call ``load_course_sheet``.
    violations != []  → reject the upload; surface to the user.

Each :class:`Violation` carries a stable ``code`` (machine-readable),
1-indexed ``row`` / ``col``, an A1 ``cell`` reference, and a human message.
The API wraps the list into a 422 response; the UI renders one row per
violation so faculty see exactly which cells to fix.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from . import config


# Allow up to 6 COs — covers the common 4 / 5 / 6 CO courses without
# letting obvious typos (e.g. a "10" or a year) sneak through silently.
MAX_CO_NUMBER: int = 6

_CO_NUMBER_RE = re.compile(r"\d+")
_SCAN_LIMIT_ROWS: int = 20  # how far down column B we look for layout markers


# ---------------------------------------------------------------------------
# Public types
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class Violation:
    """One structural problem with the uploaded workbook.

    ``code`` is stable; UI / tests should match on it. ``message`` may
    evolve. ``row`` / ``col`` are 1-indexed; ``cell`` is the A1 reference
    (e.g. ``"C7"``) when both are known.
    """

    code: str
    message: str
    row: Optional[int] = None
    col: Optional[int] = None
    cell: Optional[str] = None

    def to_dict(self) -> dict[str, Any]:
        return {k: v for k, v in asdict(self).items() if v is not None}


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _cell_ref(row: Optional[int], col: Optional[int]) -> Optional[str]:
    if row is None or col is None:
        return None
    return f"{get_column_letter(col)}{row}"


def _v(code: str, message: str,
       row: Optional[int] = None, col: Optional[int] = None) -> Violation:
    return Violation(code=code, message=message, row=row, col=col,
                     cell=_cell_ref(row, col))


def _parse_co_numbers(raw: object) -> list[int]:
    if raw is None:
        return []
    if isinstance(raw, bool):
        return []
    if isinstance(raw, (int, float)):
        return [int(raw)]
    if isinstance(raw, str):
        return [int(m.group()) for m in _CO_NUMBER_RE.finditer(raw)]
    return []


def _is_int_sl_no(value: object) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, int):
        return True
    if isinstance(value, float):
        return value.is_integer()
    return False


def _detect_layout(ws) -> dict[str, Optional[int]]:
    """Find the rows containing the four required marker labels.

    Returns a dict with keys ``question_label``, ``co_tag``, ``max_marks``,
    ``student_start``. Any value that's still ``None`` after the scan means
    the marker wasn't found — the validator surfaces these as violations.
    """
    rows: dict[str, Optional[int]] = {
        "question_label": None,
        "co_tag": None,
        "max_marks": None,
        "student_start": None,
    }
    scan = min(ws.max_row, _SCAN_LIMIT_ROWS)
    for r in range(1, scan + 1):
        b = ws.cell(row=r, column=2).value
        if b is None:
            continue
        s = str(b).strip().lower()
        if rows["question_label"] is None and s in ("question no.", "question no"):
            rows["question_label"] = r
        elif rows["co_tag"] is None and s in ("co's", "cos", "co"):
            rows["co_tag"] = r
        elif rows["max_marks"] is None and (
            s.startswith("maximum marks") or s.startswith("max marks")
        ):
            rows["max_marks"] = r
        elif rows["student_start"] is None and s == "usn":
            rows["student_start"] = r + 1
    # Also accept a "Sl.No" header in column A if column B's "USN" is missing.
    if rows["student_start"] is None:
        for r in range(1, scan + 1):
            a = ws.cell(row=r, column=1).value
            if a is None:
                continue
            if str(a).strip().lower() in ("sl.no", "sl. no", "sl no"):
                rows["student_start"] = r + 1
                break
    return rows


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def validate(path: str | Path) -> list[Violation]:
    """Return every template-contract violation in the workbook at *path*.

    A non-empty result means the upload should be rejected. The list is in
    roughly top-to-bottom, left-to-right order so the UI can show issues in
    the order a human would scan the sheet.
    """
    path = Path(path)
    violations: list[Violation] = []

    # 1. Workbook readable -------------------------------------------------
    try:
        wb = load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:
        return [_v("unreadable_file", f"Could not open workbook: {exc}")]

    # 2. Required sheet present --------------------------------------------
    sheet_name: Optional[str] = None
    if config.SHEET_CIE_SEE in wb.sheetnames:
        sheet_name = config.SHEET_CIE_SEE
    else:
        for fallback in config.SHEET_FALLBACKS:
            if fallback in wb.sheetnames:
                sheet_name = fallback
                break
    if sheet_name is None:
        return [_v(
            "missing_sheet",
            f"Workbook is missing sheet {config.SHEET_CIE_SEE!r} "
            f"(also accepted: {list(config.SHEET_FALLBACKS)}). "
            f"Found: {wb.sheetnames}",
        )]

    ws = wb[sheet_name]

    # 3. Required marker rows present --------------------------------------
    layout = _detect_layout(ws)
    marker_problems: list[Violation] = []
    if layout["question_label"] is None:
        marker_problems.append(_v(
            "missing_question_marker",
            "Could not find row labelled 'Question no.' in column B. The "
            "canonical template puts it in B5 — re-upload using the blank "
            "template if the labels were edited.",
        ))
    if layout["co_tag"] is None:
        marker_problems.append(_v(
            "missing_co_marker",
            "Could not find row labelled \"CO's\" in column B. The canonical "
            "template puts it in B6.",
        ))
    if layout["max_marks"] is None:
        marker_problems.append(_v(
            "missing_max_marks_marker",
            "Could not find row labelled 'Maximum Marks' in column B. The "
            "canonical template puts it in B7.",
        ))
    if layout["student_start"] is None:
        marker_problems.append(_v(
            "missing_student_header",
            "Could not find a 'Sl.No' / 'USN' header to anchor the student "
            "rows. The canonical template puts the header in row 9 and the "
            "first student in row 10.",
        ))
    if marker_problems:
        # Without markers we can't validate the rest reliably — return early.
        return marker_problems

    rl: int = layout["question_label"]  # type: ignore[assignment]
    rc: int = layout["co_tag"]  # type: ignore[assignment]
    rm: int = layout["max_marks"]  # type: ignore[assignment]
    rs: int = layout["student_start"]  # type: ignore[assignment]

    # 4. Walk header columns: questions, IAs, special columns --------------
    ia_index = 1
    questions_in_current_ia = 0
    special_seen = {
        config.LABEL_TEST: False,
        config.LABEL_AAT: False,
        config.LABEL_FINAL: False,
        config.LABEL_SEE: False,
    }

    for col_idx in range(config.COL_FIRST_QUESTION, ws.max_column + 1):
        label_cell = ws.cell(row=rl, column=col_idx).value
        label = (
            str(label_cell).strip()
            if label_cell is not None and str(label_cell).strip()
            else None
        )

        max_raw = ws.cell(row=rm, column=col_idx).value
        try:
            max_marks = float(max_raw) if max_raw is not None else 0.0
        except (TypeError, ValueError):
            max_marks = 0.0
            violations.append(_v(
                "non_numeric_max_marks",
                f"Max marks at {_cell_ref(rm, col_idx)} is {max_raw!r}; "
                f"expected a number.",
                row=rm, col=col_idx,
            ))

        co_raw = ws.cell(row=rc, column=col_idx).value
        co_tags = _parse_co_numbers(co_raw)
        for n in co_tags:
            if n < 1 or n > MAX_CO_NUMBER:
                violations.append(_v(
                    "co_out_of_range",
                    f"CO tag {n} at {_cell_ref(rc, col_idx)} is out of range "
                    f"[1, {MAX_CO_NUMBER}].",
                    row=rc, col=col_idx,
                ))

        if label in special_seen:
            special_seen[label] = True
            continue

        if label == config.LABEL_TOT:
            if questions_in_current_ia == 0:
                violations.append(_v(
                    "tot_without_questions",
                    f"TOT column at {_cell_ref(rl, col_idx)} but no question "
                    f"columns precede it in IA{ia_index}.",
                    row=rl, col=col_idx,
                ))
            ia_index += 1
            questions_in_current_ia = 0
            continue

        if label is not None:
            # A labelled question column (Q1, Q2, ...).
            questions_in_current_ia += 1
            if max_marks <= 0:
                violations.append(_v(
                    "missing_max_marks",
                    f"Question {label} at {_cell_ref(rl, col_idx)} has no max "
                    f"marks at {_cell_ref(rm, col_idx)}.",
                    row=rm, col=col_idx,
                ))
            if not co_tags:
                violations.append(_v(
                    "missing_co_tag",
                    f"Question {label} at {_cell_ref(rl, col_idx)} has no CO "
                    f"tag at {_cell_ref(rc, col_idx)}.",
                    row=rc, col=col_idx,
                ))
            continue

        # Unlabelled column. The parser treats it as a sub-question only if
        # both max_marks > 0 AND co_tags is non-empty. Asymmetric cells are
        # silent bugs in the parser — flag them. Fully-empty cells are
        # legitimate trailing space, so we ignore those.
        has_marks = max_marks > 0
        has_tags = bool(co_tags)
        if has_marks and not has_tags:
            violations.append(_v(
                "subcolumn_missing_co_tag",
                f"Sub-column at {_cell_ref(rl, col_idx)} has max marks "
                f"{max_marks!r} but no CO tag at {_cell_ref(rc, col_idx)}; "
                f"the calculator would silently drop it.",
                row=rc, col=col_idx,
            ))
        elif has_tags and not has_marks:
            violations.append(_v(
                "subcolumn_missing_max_marks",
                f"Sub-column at {_cell_ref(rl, col_idx)} has CO tag {co_tags!r} "
                f"but no max marks at {_cell_ref(rm, col_idx)}; the "
                f"calculator would silently drop it.",
                row=rm, col=col_idx,
            ))

    # 5. Required structural elements --------------------------------------
    if ia_index == 1:
        violations.append(_v(
            "no_ia_blocks",
            f"No '{config.LABEL_TOT}' columns found — the workbook does not "
            f"contain any IA blocks.",
            row=rl,
        ))
    if not special_seen[config.LABEL_SEE]:
        violations.append(_v(
            "missing_see_column",
            f"No '{config.LABEL_SEE}' column found in the labels row "
            f"({rl}).",
            row=rl,
        ))
    if not special_seen[config.LABEL_AAT]:
        violations.append(_v(
            "missing_aat_column",
            f"No '{config.LABEL_AAT}' column found in the labels row "
            f"({rl}).",
            row=rl,
        ))

    # 6. Student rows ------------------------------------------------------
    seen_student = False
    range_ended = False
    for row in range(rs, ws.max_row + 1):
        a = ws.cell(row=row, column=1).value
        b = ws.cell(row=row, column=2).value
        a_blank = a is None or (isinstance(a, str) and not a.strip())
        b_blank = b is None or (isinstance(b, str) and not str(b).strip())

        if _is_int_sl_no(a):
            if range_ended:
                violations.append(_v(
                    "student_after_summary",
                    f"Row {row} has integer Sl.No {int(a)} but the student "
                    f"range already ended earlier — student rows must be "
                    f"contiguous.",
                    row=row, col=1,
                ))
                continue
            seen_student = True
            if b_blank:
                violations.append(_v(
                    "missing_usn",
                    f"Row {row} has Sl.No {int(a)} but no USN at column B.",
                    row=row, col=2,
                ))
            continue

        if a_blank and b_blank:
            # Fully blank row — ignored both inside and after the student
            # range, matching the parser's behavior.
            continue

        # Non-int Sl.No (or non-blank col B with blank col A). The parser
        # stops here. If we've already seen students, this is the legitimate
        # summary; otherwise the student range never started, which is a
        # real bug.
        if seen_student:
            range_ended = True
        else:
            violations.append(_v(
                "non_int_sl_no",
                f"Row {row} Sl.No is {a!r}; expected an integer or a blank "
                f"cell.",
                row=row, col=1,
            ))
            break

    if not seen_student:
        violations.append(_v(
            "no_students",
            f"No integer Sl.No found from row {rs} onward — the sheet "
            f"contains no student rows.",
            row=rs, col=1,
        ))

    return violations
