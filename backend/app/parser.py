"""Read a faculty CIE+SEE workbook into a :class:`CourseSheet`.

The parser is intentionally tolerant — CO tags can be ints, comma-lists, or
phrases like ``"CO 1 2 3 4"``; blank student cells mean "not attempted".
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from . import config
from .models import Column, ColumnKind, CourseSheet, StudentRow


_CO_NUMBER_RE = re.compile(r"\d+")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_co_tags(raw: object) -> tuple[int, ...]:
    """Extract the integer CO numbers from a row-6 tag cell.

    Handles:
      - ``None``                   -> ``()``
      - int ``2``                  -> ``(2,)``
      - str ``"2,3"`` / ``"2 3"``  -> ``(2, 3)``
      - str ``"CO 1 2 3 4"``       -> ``(1, 2, 3, 4)``
    """
    if raw is None:
        return ()
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return (int(raw),)
    if isinstance(raw, str):
        nums = [int(m.group()) for m in _CO_NUMBER_RE.finditer(raw)]
        # In strings like "CO 1 2 3 4" the leading "CO" contributes no digits,
        # so this works without special casing.
        return tuple(nums)
    return ()


def _parse_mark(cell: Cell) -> float | str | None:
    """Return ``None`` for a truly blank cell, ``float`` for numeric marks,
    or the raw ``str`` (e.g. ``"NE"``) for non-numeric text cells.

    Matching Excel semantics: ``COUNTIF(range, "<>")`` counts any non-blank
    cell including text, so we must distinguish ``None`` from ``"NE"``.
    """
    value = cell.value
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.strip())
        except ValueError:
            return value.strip()  # e.g. "NE" — present but not numeric
    return None


_SPECIAL_LABELS = {
    config.LABEL_TEST: "test",
    config.LABEL_AAT: "aat",
    config.LABEL_FINAL: "final",
    config.LABEL_SEE: "see",
}


def _classify_column(
    label: str | None,
    max_marks: float,
    co_tags: tuple[int, ...],
    current_question_label: str | None,
    ia_counter: dict[str, int],
) -> tuple[ColumnKind | None, int | None, str | None]:
    """Return ``(kind, ia_index, resolved_label)`` for one column.

    Layout quirk: each CIE question spans **two** sub-columns. Only the first
    has a row-5 label (``"Q1"``) — the second is unlabeled but carries its
    own CO tag and max marks. We treat both as separate "question" columns,
    with the unlabeled one inheriting the label (``"Q1"``) of its parent.

    Returns ``(None, None, None)`` for columns that should be skipped.
    """
    if label in _SPECIAL_LABELS:
        return _SPECIAL_LABELS[label], None, label
    if label == config.LABEL_TOT:
        ia = ia_counter["current"]
        ia_counter["current"] += 1  # next block of Qs belongs to the next IA
        return "tot", ia, label
    # A labelled question column ("Q1", "Q2", ...).
    if label is not None:
        return "question", ia_counter["current"], label
    # An unlabeled sub-column — treat as a question sub-part **only if** it
    # has a max mark and a CO tag (otherwise it's just trailing empty space).
    if max_marks > 0 and co_tags:
        return "question", ia_counter["current"], current_question_label
    return None, None, None


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def _find_sheet(wb) -> str:
    """Return the name of the CIE+SEE (or equivalent) sheet."""
    if config.SHEET_CIE_SEE in wb.sheetnames:
        return config.SHEET_CIE_SEE
    for name in config.SHEET_FALLBACKS:
        if name in wb.sheetnames:
            return name
    raise ValueError(
        f"Workbook is missing required sheet "
        f"{config.SHEET_CIE_SEE!r}; found {wb.sheetnames}"
    )


def _detect_layout(ws) -> dict[str, int]:
    """Auto-detect key row numbers by scanning for marker cells.

    Returns a dict with keys: question_label, co_tag, max_marks, student_start.
    """
    row_question_label = config.ROW_QUESTION_LABEL
    row_co_tag = config.ROW_CO_TAG
    row_max_marks = config.ROW_MAX_MARKS
    row_student_start = config.ROW_STUDENT_START

    scan_limit = min(ws.max_row, 20)
    for r in range(1, scan_limit + 1):
        col_b = ws.cell(row=r, column=2).value
        if col_b is None:
            continue
        s = str(col_b).strip().lower()
        if s in ("question no.", "question no"):
            row_question_label = r
        elif s in ("co's", "cos", "co"):
            row_co_tag = r
        elif s.startswith("maximum marks") or s.startswith("max marks"):
            row_max_marks = r
        elif s in ("usn",):
            # The header row with "Sl.No" / "USN" — students start next row.
            row_student_start = r + 1

    # Also check col A for "Sl.No" marker.
    for r in range(1, scan_limit + 1):
        col_a = ws.cell(row=r, column=1).value
        if col_a is not None and str(col_a).strip().lower() in ("sl.no", "sl. no", "sl no"):
            row_student_start = r + 1
            break

    return {
        "question_label": row_question_label,
        "co_tag": row_co_tag,
        "max_marks": row_max_marks,
        "student_start": row_student_start,
    }


def load_course_sheet(path: str | Path) -> CourseSheet:
    """Parse the CIE+SEE sheet of *path* into a :class:`CourseSheet`."""
    path = Path(path)
    wb = load_workbook(path, data_only=True, read_only=False)
    sheet_name = _find_sheet(wb)
    ws = wb[sheet_name]

    # Auto-detect row layout.
    layout = _detect_layout(ws)
    row_question_label = layout["question_label"]
    row_co_tag = layout["co_tag"]
    row_max_marks = layout["max_marks"]
    row_student_start = layout["student_start"]

    course_name = ""
    # Try "SUB" label in row 2 col A (canonical template).
    sub_label = ws.cell(row=2, column=1).value
    sub_value = ws.cell(row=2, column=2).value
    if isinstance(sub_label, str) and sub_label.strip().upper() == "SUB" and sub_value:
        course_name = str(sub_value).strip()

    # --- Columns --------------------------------------------------------------
    columns: list[Column] = []
    co_numbers_seen: set[int] = set()
    ia_counter = {"current": 1}
    last_question_label: str | None = None

    for col_idx in range(config.COL_FIRST_QUESTION, ws.max_column + 1):
        label_cell = ws.cell(row=row_question_label, column=col_idx).value
        label = (
            str(label_cell).strip()
            if label_cell is not None and str(label_cell).strip()
            else None
        )

        max_marks_raw = ws.cell(row=row_max_marks, column=col_idx).value
        try:
            max_marks = float(max_marks_raw) if max_marks_raw is not None else 0.0
        except (TypeError, ValueError):
            max_marks = 0.0

        co_tags = _parse_co_tags(ws.cell(row=row_co_tag, column=col_idx).value)

        kind, ia_index, resolved_label = _classify_column(
            label, max_marks, co_tags, last_question_label, ia_counter,
        )
        if kind is None:
            continue
        if kind == "question":
            if label is not None:
                last_question_label = label

        co_numbers_seen.update(co_tags)
        columns.append(Column(
            index=col_idx,
            kind=kind,
            label=resolved_label or "",
            max_marks=max_marks,
            co_tags=co_tags,
            ia_index=ia_index,
        ))

    # --- Students -------------------------------------------------------------
    students: list[StudentRow] = []
    for row in range(row_student_start, ws.max_row + 1):
        sl_no_cell = ws.cell(row=row, column=config.COL_SL_NO).value
        usn_cell = ws.cell(row=row, column=config.COL_USN).value
        if not isinstance(sl_no_cell, (int, float)) or isinstance(sl_no_cell, bool):
            if sl_no_cell is None and (usn_cell is None or not str(usn_cell).strip()):
                continue
            break
        sl_no = int(sl_no_cell)

        marks: dict[int, float | None] = {}
        for col in columns:
            marks[col.index] = _parse_mark(ws.cell(row=row, column=col.index))
        students.append(StudentRow(
            sl_no=sl_no,
            usn=str(usn_cell).strip() if usn_cell else "",
            marks=marks,
        ))

    co_numbers = tuple(sorted(co_numbers_seen))
    return CourseSheet(
        path=str(path),
        course_name=course_name,
        columns=tuple(columns),
        students=tuple(students),
        co_numbers=co_numbers,
    )
