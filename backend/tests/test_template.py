"""Unit tests for the template validator.

We build small synthetic workbooks in-memory (no fixture files on disk) so
each test pins a single, named contract violation. The real-faculty sample
is also covered: it must validate clean.
"""
from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app import config
from app.template import MAX_CO_NUMBER, Violation, validate


SAMPLE = Path(__file__).resolve().parents[1] / "samples" / "python_programming_21ET641.xlsx"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _codes(violations: list[Violation]) -> list[str]:
    return [v.code for v in violations]


def _make_minimal_workbook() -> Workbook:
    """A 1-IA, 1-Q, 1-student CIE+SEE workbook that should pass validation.

    Layout is intentionally compact:
      - B5 'Question no.', B6 "CO's", B7 'Maximum Marks', B9 'USN'
      - C5 'Q1', C6 1, C7 10
      - D5 'TOT', D7 10        (terminates IA1)
      - E5 'AAT (ASSGN)', E6 1, E7 20
      - F5 'SEE', F6 1, F7 100
      - A10 1, B10 'X'         (one student)
    Anything beyond that is for tests to mutate.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = config.SHEET_CIE_SEE
    ws["B5"] = "Question no."
    ws["B6"] = "CO's"
    ws["B7"] = "Maximum Marks"
    ws["B9"] = "USN"
    ws["A9"] = "Sl.No"
    # Q1 (CO1, 10 marks)
    ws["C5"] = "Q1"
    ws["C6"] = 1
    ws["C7"] = 10
    # TOT
    ws["D5"] = "TOT"
    ws["D7"] = 10
    # AAT
    ws["E5"] = "AAT (ASSGN)"
    ws["E6"] = "CO 1"
    ws["E7"] = 20
    # SEE
    ws["F5"] = "SEE"
    ws["F6"] = 1
    ws["F7"] = 100
    # One student
    ws["A10"] = 1
    ws["B10"] = "1DSTEST"
    return wb


def _save(wb: Workbook, tmp_path: Path, name: str = "wb.xlsx") -> Path:
    p = tmp_path / name
    wb.save(p)
    return p


# ---------------------------------------------------------------------------
# Happy paths
# ---------------------------------------------------------------------------


@pytest.mark.skipif(not SAMPLE.exists(), reason=f"sample missing at {SAMPLE}")
def test_canonical_sample_passes_clean():
    """The real faculty workbook must validate with zero violations."""
    assert validate(SAMPLE) == []


def test_minimal_synthetic_workbook_passes(tmp_path):
    p = _save(_make_minimal_workbook(), tmp_path)
    assert validate(p) == []


# ---------------------------------------------------------------------------
# File / sheet-level failures
# ---------------------------------------------------------------------------


def test_unreadable_file(tmp_path):
    bad = tmp_path / "garbage.xlsx"
    bad.write_bytes(b"this is not a real xlsx file")
    vs = validate(bad)
    assert _codes(vs) == ["unreadable_file"]


def test_missing_sheet(tmp_path):
    wb = Workbook()
    wb.active.title = "Something Else"
    p = _save(wb, tmp_path)
    vs = validate(p)
    assert _codes(vs) == ["missing_sheet"]
    assert "CIE+SEE" in vs[0].message


def test_sheet1_fallback_accepted(tmp_path):
    """If 'CIE+SEE' is missing but 'Sheet1' is present, it's accepted."""
    wb = _make_minimal_workbook()
    wb.active.title = "Sheet1"
    p = _save(wb, tmp_path)
    assert validate(p) == []


# ---------------------------------------------------------------------------
# Marker-row failures (return early — no other violations reported)
# ---------------------------------------------------------------------------


def test_missing_question_marker(tmp_path):
    wb = _make_minimal_workbook()
    wb.active["B5"] = None
    p = _save(wb, tmp_path)
    assert _codes(validate(p)) == ["missing_question_marker"]


def test_missing_co_marker(tmp_path):
    wb = _make_minimal_workbook()
    wb.active["B6"] = None
    p = _save(wb, tmp_path)
    assert _codes(validate(p)) == ["missing_co_marker"]


def test_missing_max_marks_marker(tmp_path):
    wb = _make_minimal_workbook()
    wb.active["B7"] = None
    p = _save(wb, tmp_path)
    assert _codes(validate(p)) == ["missing_max_marks_marker"]


def test_missing_student_header(tmp_path):
    wb = _make_minimal_workbook()
    wb.active["B9"] = None
    wb.active["A9"] = None
    p = _save(wb, tmp_path)
    assert _codes(validate(p)) == ["missing_student_header"]


def test_marker_at_non_default_row_is_accepted(tmp_path):
    """Markers may live at any row in the first 20 — we auto-detect."""
    wb = Workbook()
    ws = wb.active
    ws.title = config.SHEET_CIE_SEE
    # Same content as the minimal workbook but shifted down by 2 rows.
    ws["B7"] = "Question no."
    ws["B8"] = "CO's"
    ws["B9"] = "Maximum Marks"
    ws["B11"] = "USN"
    ws["A11"] = "Sl.No"
    ws["C7"] = "Q1"
    ws["C8"] = 1
    ws["C9"] = 10
    ws["D7"] = "TOT"
    ws["D9"] = 10
    ws["E7"] = "AAT (ASSGN)"
    ws["E8"] = "CO 1"
    ws["E9"] = 20
    ws["F7"] = "SEE"
    ws["F8"] = 1
    ws["F9"] = 100
    ws["A12"] = 1
    ws["B12"] = "1DSTEST"
    p = _save(wb, tmp_path)
    assert validate(p) == []


# ---------------------------------------------------------------------------
# Column-level failures
# ---------------------------------------------------------------------------


def test_missing_co_tag_on_question(tmp_path):
    wb = _make_minimal_workbook()
    wb.active["C6"] = None
    p = _save(wb, tmp_path)
    codes = _codes(validate(p))
    assert "missing_co_tag" in codes


def test_missing_max_marks_on_question(tmp_path):
    wb = _make_minimal_workbook()
    wb.active["C7"] = None
    p = _save(wb, tmp_path)
    codes = _codes(validate(p))
    assert "missing_max_marks" in codes


def test_non_numeric_max_marks(tmp_path):
    wb = _make_minimal_workbook()
    wb.active["C7"] = "ten"
    p = _save(wb, tmp_path)
    codes = _codes(validate(p))
    # "ten" → non_numeric_max_marks AND (since max_marks parses as 0)
    # missing_max_marks. Both are accurate diagnoses.
    assert "non_numeric_max_marks" in codes


def test_co_out_of_range(tmp_path):
    wb = _make_minimal_workbook()
    wb.active["C6"] = MAX_CO_NUMBER + 1
    p = _save(wb, tmp_path)
    vs = validate(p)
    assert "co_out_of_range" in _codes(vs)
    # Locate the violation and check its cell reference.
    out = next(v for v in vs if v.code == "co_out_of_range")
    assert out.cell == "C6"


def test_subcolumn_missing_co_tag(tmp_path):
    """Unlabelled sub-column with max_marks but no CO tag is the silent
    drop the parser is famous for. Validator must catch."""
    wb = _make_minimal_workbook()
    # Insert a sub-column at D (between Q1 in C and TOT in original D).
    # We need to shift the rest right; simpler: rebuild with the bad shape.
    ws = wb.active
    ws["D5"] = None       # unlabelled
    ws["D6"] = None       # no CO tag
    ws["D7"] = 5          # has max marks → asymmetric
    # TOT moves to E, AAT to F, SEE to G.
    ws["E5"] = "TOT"
    ws["E7"] = 10
    ws["F5"] = "AAT (ASSGN)"
    ws["F6"] = "CO 1"
    ws["F7"] = 20
    ws["G5"] = "SEE"
    ws["G6"] = 1
    ws["G7"] = 100
    p = _save(wb, tmp_path)
    codes = _codes(validate(p))
    assert "subcolumn_missing_co_tag" in codes


def test_subcolumn_missing_max_marks(tmp_path):
    wb = _make_minimal_workbook()
    ws = wb.active
    ws["D5"] = None       # unlabelled
    ws["D6"] = 1          # has CO tag
    ws["D7"] = None       # but no max marks → asymmetric
    ws["E5"] = "TOT"
    ws["E7"] = 10
    ws["F5"] = "AAT (ASSGN)"
    ws["F6"] = "CO 1"
    ws["F7"] = 20
    ws["G5"] = "SEE"
    ws["G6"] = 1
    ws["G7"] = 100
    p = _save(wb, tmp_path)
    codes = _codes(validate(p))
    assert "subcolumn_missing_max_marks" in codes


def test_pure_empty_subcolumn_is_not_flagged(tmp_path):
    """A column with no label, no marks, no CO tag is just trailing space —
    the parser ignores it and the validator must not complain."""
    wb = _make_minimal_workbook()
    ws = wb.active
    # Insert a fully-empty column at D, shifting TOT/AAT/SEE right.
    ws["D5"] = None
    ws["D6"] = None
    ws["D7"] = None
    ws["E5"] = "TOT"
    ws["E7"] = 10
    ws["F5"] = "AAT (ASSGN)"
    ws["F6"] = "CO 1"
    ws["F7"] = 20
    ws["G5"] = "SEE"
    ws["G6"] = 1
    ws["G7"] = 100
    p = _save(wb, tmp_path)
    assert validate(p) == []


# ---------------------------------------------------------------------------
# Structural / IA-level failures
# ---------------------------------------------------------------------------


def test_no_ia_blocks(tmp_path):
    wb = _make_minimal_workbook()
    wb.active["D5"] = None  # delete the only TOT
    wb.active["D7"] = None
    p = _save(wb, tmp_path)
    codes = _codes(validate(p))
    assert "no_ia_blocks" in codes


def test_missing_see_column(tmp_path):
    wb = _make_minimal_workbook()
    wb.active["F5"] = None  # delete SEE label
    p = _save(wb, tmp_path)
    codes = _codes(validate(p))
    assert "missing_see_column" in codes


def test_missing_aat_column(tmp_path):
    wb = _make_minimal_workbook()
    wb.active["E5"] = None  # delete AAT label
    p = _save(wb, tmp_path)
    codes = _codes(validate(p))
    assert "missing_aat_column" in codes


def test_tot_without_questions(tmp_path):
    """A TOT immediately after another TOT (or at the very start of the IA
    section) is a structural mistake — no questions to total."""
    wb = _make_minimal_workbook()
    ws = wb.active
    # Replace Q1 column with a TOT: now we have TOT at C, TOT at D, AAT, SEE.
    ws["C5"] = "TOT"
    ws["C6"] = None
    ws["C7"] = 10
    p = _save(wb, tmp_path)
    codes = _codes(validate(p))
    assert "tot_without_questions" in codes


# ---------------------------------------------------------------------------
# Student-row failures
# ---------------------------------------------------------------------------


def test_no_students(tmp_path):
    wb = _make_minimal_workbook()
    wb.active["A10"] = None
    wb.active["B10"] = None
    p = _save(wb, tmp_path)
    codes = _codes(validate(p))
    assert "no_students" in codes


def test_non_int_sl_no_at_start(tmp_path):
    wb = _make_minimal_workbook()
    wb.active["A10"] = "one"
    p = _save(wb, tmp_path)
    codes = _codes(validate(p))
    assert "non_int_sl_no" in codes


def test_missing_usn(tmp_path):
    wb = _make_minimal_workbook()
    wb.active["B10"] = None
    p = _save(wb, tmp_path)
    codes = _codes(validate(p))
    assert "missing_usn" in codes


def test_student_after_summary(tmp_path):
    """Integer Sl.No after the summary section started is a contiguity bug."""
    wb = _make_minimal_workbook()
    ws = wb.active
    ws["A10"] = 1
    ws["B10"] = "1DSTEST1"
    ws["A11"] = 2
    ws["B11"] = "1DSTEST2"
    ws["A12"] = None
    ws["B12"] = "count >60%"   # summary row — ends the student range
    ws["A13"] = 3              # but here's another integer Sl.No → bug
    ws["B13"] = "1DSTEST3"
    p = _save(wb, tmp_path)
    codes = _codes(validate(p))
    assert "student_after_summary" in codes


def test_summary_rows_after_students_are_fine(tmp_path):
    """Pure summary rows (col A blank, col B = 'count >60%' etc.) end the
    student range without any violation."""
    wb = _make_minimal_workbook()
    ws = wb.active
    ws["A11"] = None
    ws["B11"] = "count >60%"
    ws["A12"] = None
    ws["B12"] = "count of attended"
    ws["A13"] = None
    ws["B13"] = "CO Attainment = B/A"
    p = _save(wb, tmp_path)
    assert validate(p) == []


def test_blank_rows_inside_student_range_are_fine(tmp_path):
    wb = _make_minimal_workbook()
    ws = wb.active
    ws["A10"] = 1
    ws["B10"] = "1DSTEST1"
    # blank row 11
    ws["A12"] = 2
    ws["B12"] = "1DSTEST2"
    p = _save(wb, tmp_path)
    assert validate(p) == []


# ---------------------------------------------------------------------------
# Violation serialisation (used by the API → UI)
# ---------------------------------------------------------------------------


def test_violation_to_dict_roundtrip():
    v = Violation(code="x", message="msg", row=5, col=3, cell="C5")
    d = v.to_dict()
    assert d == {"code": "x", "message": "msg", "row": 5, "col": 3, "cell": "C5"}


def test_violation_to_dict_drops_none_fields():
    v = Violation(code="missing_sheet", message="…")
    d = v.to_dict()
    assert d == {"code": "missing_sheet", "message": "…"}
    assert "row" not in d
