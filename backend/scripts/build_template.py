"""Build ``backend/samples/template_v1.xlsx`` from the canonical sample.

Produces a single-sheet, course-agnostic blank: only ``CIE+SEE`` is kept,
student data and computed summary numbers are cleared, and the per-question
max-marks row is blanked so faculty fill it in for their own course.

Run from the repo root::

    python backend/scripts/build_template.py

The resulting file is committed to the repo and served by
``GET /api/v1/template``.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

# Make ``app.*`` importable when running this as a script.
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app import config


SAMPLE = Path(__file__).resolve().parents[1] / "samples" / "python_programming_21ET641.xlsx"
OUT = Path(__file__).resolve().parents[1] / "samples" / "template_v1.xlsx"


def _is_int(v: object) -> bool:
    if isinstance(v, bool):
        return False
    if isinstance(v, int):
        return True
    if isinstance(v, float):
        return v.is_integer()
    return False


def _clear_cell(ws, row: int, col: int) -> None:
    """Set a cell to None, skipping merged-range tail cells (read-only)."""
    cell = ws.cell(row=row, column=col)
    if cell.__class__.__name__ == "MergedCell":
        return
    cell.value = None


def _is_formula(v: object) -> bool:
    return isinstance(v, str) and v.startswith("=")


def main() -> None:
    if not SAMPLE.exists():
        raise SystemExit(f"sample missing at {SAMPLE}")

    wb = load_workbook(SAMPLE)

    # --- Drop every sheet except CIE+SEE so the template is one page.
    for name in list(wb.sheetnames):
        if name != config.SHEET_CIE_SEE:
            del wb[name]

    ws = wb[config.SHEET_CIE_SEE]

    # --- Blank the per-question max-marks row, but keep the TOT SUM
    # formulas so the IA totals auto-fill once faculty enter the marks.
    cleared_max_marks = 0
    for col in range(config.COL_FIRST_QUESTION, ws.max_column + 1):
        cell = ws.cell(row=config.ROW_MAX_MARKS, column=col)
        if cell.__class__.__name__ == "MergedCell":
            continue
        if _is_formula(cell.value):
            continue
        if cell.value is not None:
            cell.value = None
            cleared_max_marks += 1

    # --- Student rows + computed summary rows.
    cleared_students = 0
    cleared_summary = 0
    for row in range(config.ROW_STUDENT_START, ws.max_row + 1):
        a = ws.cell(row=row, column=1).value
        if _is_int(a):
            # Student row — wipe everything from A through max_col.
            for col in range(1, ws.max_column + 1):
                _clear_cell(ws, row, col)
            cleared_students += 1
        else:
            # Summary row (count >60%, count attended, CO Attainment, etc.).
            # Wipe the right-hand side so faculty don't ship stale numbers,
            # but keep col B's row labels intact.
            for col in range(3, ws.max_column + 1):
                _clear_cell(ws, row, col)
            cleared_summary += 1

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(
        f"wrote {OUT}  (kept only '{config.SHEET_CIE_SEE}'; "
        f"cleared {cleared_max_marks} max-marks cells, "
        f"{cleared_students} student rows, {cleared_summary} summary rows)"
    )


if __name__ == "__main__":
    main()
